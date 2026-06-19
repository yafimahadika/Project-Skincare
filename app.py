from functools import wraps
import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from math import log2

from flask import Flask, Response, flash, redirect, render_template, request, session, url_for
from flask_mysqldb import MySQL
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from config import Config


app = Flask(__name__)
app.config.from_object(Config)
mysql = MySQL(app)


@app.template_filter("format_date_id")
def format_date_id(value):
    if not value:
        return "-"

    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    try:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return str(value)


@app.template_filter("format_rupiah")
def format_rupiah(value):
    try:
        return "Rp " + "{:,.0f}".format(float(value or 0)).replace(",", ".")
    except (TypeError, ValueError):
        return "Rp 0"


@app.template_filter("format_period")
def format_period(tanggal_awal, tanggal_akhir):
    return f"{format_date_id(tanggal_awal)} - {format_date_id(tanggal_akhir)}"


def normalize_date_input(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    value = (value or "").strip()
    if not value:
        return ""

    for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, date_format).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return ""


def get_period_inputs():
    tanggal_awal = normalize_date_input(request.values.get("tanggal_awal", ""))
    tanggal_akhir = normalize_date_input(request.values.get("tanggal_akhir", ""))
    return tanggal_awal, tanggal_akhir


def parse_int(value, default=0):
    try:
        return int(float(value or default))
    except (TypeError, ValueError):
        return default


def parse_float(value, default=0):
    try:
        return float(value or default)
    except (TypeError, ValueError):
        return default


def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if "user_id" not in session:
            flash("Silakan login terlebih dahulu.", "warning")
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped_view


def admin_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Menu ini hanya dapat diakses oleh Admin.", "danger")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped_view


def calculate_entropy(rows):
    total = len(rows)
    if total == 0:
        return 0

    class_counts = {}
    for row in rows:
        label = row.get("label_klasifikasi") or "-"
        class_counts[label] = class_counts.get(label, 0) + 1

    entropy = 0
    for count in class_counts.values():
        probability = count / total
        if probability > 0:
            entropy -= probability * log2(probability)

    return entropy


def get_system_label(jumlah_terjual, average_sold):
    return "Laris" if float(jumlah_terjual or 0) >= average_sold else "Tidak Laris"


def categorize_number(value, low_limit, high_limit):
    value = float(value or 0)
    if value <= low_limit:
        return "Rendah"
    if value <= high_limit:
        return "Sedang"
    return "Tinggi"


def prepare_c45_dataset(rows):
    if not rows:
        return []

    average_sold = sum(float(row.get("jumlah_terjual") or 0) for row in rows) / len(rows)
    numeric_fields = ["harga", "jumlah_terjual", "total_penjualan"]
    limits = {}

    for field in numeric_fields:
        values = sorted(float(row.get(field) or 0) for row in rows)
        low_index = max(0, int((len(values) - 1) * 0.33))
        high_index = max(0, int((len(values) - 1) * 0.66))
        limits[field] = (values[low_index], values[high_index])

    dataset = []
    for row in rows:
        dataset.append(
            {
                "kategori": row.get("kategori") or "-",
                "bulan": row.get("bulan") or "-",
                "harga": categorize_number(row.get("harga"), *limits["harga"]),
                "jumlah_terjual": categorize_number(row.get("jumlah_terjual"), *limits["jumlah_terjual"]),
                "total_penjualan": categorize_number(row.get("total_penjualan"), *limits["total_penjualan"]),
                "label_klasifikasi": get_system_label(row.get("jumlah_terjual"), average_sold),
            }
        )

    return dataset


def parse_sales_rows(uploaded_file):
    filename = secure_filename(uploaded_file.filename or "")
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    rows = []
    if extension == "csv":
        content = uploaded_file.stream.read().decode("utf-8-sig")
        reader = csv.DictReader(StringIO(content))
        rows = list(reader)
    elif extension == "xlsx":
        workbook = load_workbook(uploaded_file.stream, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({headers[index]: value for index, value in enumerate(values) if index < len(headers)})
    else:
        raise ValueError("Format file harus CSV atau XLSX.")

    return rows


def month_name_id(month_number):
    months = [
        "Januari",
        "Februari",
        "Maret",
        "April",
        "Mei",
        "Juni",
        "Juli",
        "Agustus",
        "September",
        "Oktober",
        "November",
        "Desember",
    ]
    return months[month_number - 1]


def calculate_gain_ratio(dataset, attribute, base_entropy):
    total = len(dataset)
    groups = {}

    for row in dataset:
        value = row.get(attribute) or "-"
        groups.setdefault(value, []).append(row)

    weighted_entropy = 0
    split_info = 0
    details = []

    for value, rows in groups.items():
        ratio = len(rows) / total
        entropy_value = calculate_entropy(rows)
        weighted_entropy += ratio * entropy_value

        if ratio > 0:
            split_info -= ratio * log2(ratio)

        class_counts = {}
        for row in rows:
            label = row.get("label_klasifikasi") or "-"
            class_counts[label] = class_counts.get(label, 0) + 1

        details.append(
            {
                "value": value,
                "total": len(rows),
                "entropy": entropy_value,
                "class_counts": class_counts,
            }
        )

    gain = base_entropy - weighted_entropy
    gain_ratio = gain / split_info if split_info else 0

    return {
        "attribute": attribute,
        "gain": gain,
        "split_info": split_info,
        "gain_ratio": gain_ratio,
        "details": details,
    }


def build_simple_rules(dataset, attribute):
    rules = []
    if not attribute:
        return rules

    groups = {}
    for row in dataset:
        groups.setdefault(row.get(attribute) or "-", []).append(row)

    for value, rows in groups.items():
        class_counts = {}
        for row in rows:
            label = row.get("label_klasifikasi") or "-"
            class_counts[label] = class_counts.get(label, 0) + 1

        result = max(class_counts, key=class_counts.get)
        rules.append(
            {
                "condition": f"Jika {attribute.replace('_', ' ').title()} = {value}",
                "result": result,
                "total": len(rows),
            }
        )

    return rules


def get_c45_summary(raw_dataset):
    dataset = prepare_c45_dataset(raw_dataset)
    base_entropy = calculate_entropy(dataset)
    attributes = [
        ("kategori", "Kategori"),
        ("bulan", "Bulan"),
        ("harga", "Harga"),
        ("jumlah_terjual", "Jumlah Terjual"),
        ("total_penjualan", "Total Penjualan"),
    ]

    calculations = []
    if dataset:
        calculations = [
            {**calculate_gain_ratio(dataset, key, base_entropy), "label": label}
            for key, label in attributes
        ]
        calculations.sort(key=lambda item: item["gain_ratio"], reverse=True)

    best_attribute = calculations[0] if calculations else None
    return dataset, base_entropy, calculations, best_attribute


def get_aggregated_sales(tanggal_awal, tanggal_akhir):
    if not tanggal_awal or not tanggal_akhir:
        return []

    cursor = mysql.connection.cursor()
    cursor.execute(
        """
        SELECT
            p.id AS produk_id,
            p.kode_produk,
            p.nama_produk,
            p.kategori,
            p.harga,
            SUM(pjl.jumlah_terjual) AS jumlah_terjual,
            SUM(pjl.total_penjualan) AS total_penjualan
        FROM penjualan pjl
        JOIN produk p ON p.id = pjl.produk_id
        WHERE pjl.tanggal_penjualan BETWEEN %s AND %s
        GROUP BY p.id, p.kode_produk, p.nama_produk, p.kategori, p.harga
        ORDER BY p.nama_produk ASC
        """,
        (tanggal_awal, tanggal_akhir),
    )
    rows = cursor.fetchall()
    cursor.close()
    return rows


def build_classification_rows(raw_dataset):
    dataset, base_entropy, calculations, best_attribute = get_c45_summary(raw_dataset)
    average_sold = (
        sum(float(row.get("jumlah_terjual") or 0) for row in raw_dataset) / len(raw_dataset)
        if raw_dataset else 0
    )
    gain_ratio = best_attribute["gain_ratio"] if best_attribute else 0
    attribute_label = best_attribute["label"] if best_attribute else "-"
    rows = []

    for row in raw_dataset:
        jumlah_terjual = float(row.get("jumlah_terjual") or 0)
        hasil = get_system_label(jumlah_terjual, average_sold)
        comparison = "lebih besar atau sama dengan" if hasil == "Laris" else "berada di bawah"
        keterangan = (
            f"Produk dikategorikan {hasil} karena jumlah terjual {comparison} "
            f"rata-rata penjualan pada periode tersebut. "
            f"Atribut terbaik: {attribute_label}, gain ratio {gain_ratio:.6f}."
        )
        rows.append({**row, "rata_rata_penjualan": average_sold, "hasil_klasifikasi": hasil, "keterangan": keterangan})

    return rows, dataset, base_entropy, calculations, best_attribute


def get_classification_filters():
    tanggal_awal = normalize_date_input(request.args.get("tanggal_awal", ""))
    tanggal_akhir = normalize_date_input(request.args.get("tanggal_akhir", ""))
    filter_bulan = request.args.get("bulan", "").strip()
    filter_tahun = request.args.get("tahun", "").strip()
    filter_hasil = request.args.get("hasil", "").strip()
    return tanggal_awal, tanggal_akhir, filter_bulan, filter_tahun, filter_hasil


def fetch_classification_results():
    tanggal_awal, tanggal_akhir, filter_bulan, filter_tahun, filter_hasil = get_classification_filters()
    conditions = []
    params = []

    if tanggal_awal:
        conditions.append("tanggal_awal >= %s")
        params.append(tanggal_awal)
    if tanggal_akhir:
        conditions.append("tanggal_akhir <= %s")
        params.append(tanggal_akhir)
    if filter_bulan:
        conditions.append("bulan = %s")
        params.append(filter_bulan)
    if filter_tahun:
        conditions.append("tahun = %s")
        params.append(filter_tahun)
    if filter_hasil:
        conditions.append("hasil_klasifikasi = %s")
        params.append(filter_hasil)

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    cursor = mysql.connection.cursor()
    cursor.execute(
        f"""
        SELECT *
        FROM hasil_klasifikasi
        {where_clause}
        ORDER BY created_at DESC, id DESC
        """,
        tuple(params),
    )
    rows = cursor.fetchall()
    cursor.close()
    return rows, {
        "tanggal_awal": tanggal_awal,
        "tanggal_akhir": tanggal_akhir,
        "bulan": filter_bulan,
        "tahun": filter_tahun,
        "hasil": filter_hasil,
    }


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        cursor = mysql.connection.cursor()
        cursor.execute(
            "SELECT * FROM users WHERE username = %s",
            (username,),
        )
        user = cursor.fetchone()
        cursor.close()

        if user and check_password_hash(user["password"], password):
            session["user_id"] = user["id"]
            session["nama"] = user["nama"]
            session["role"] = user["role"]
            flash("Login berhasil.", "success")
            return redirect(url_for("dashboard"))

        flash("Username atau password salah.", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM produk")
    total_produk = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM penjualan")
    total_penjualan = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM hasil_klasifikasi")
    total_klasifikasi = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM hasil_klasifikasi WHERE hasil_klasifikasi = 'Laris'")
    total_laris = cursor.fetchone()["total"]
    cursor.execute("SELECT COUNT(*) AS total FROM hasil_klasifikasi WHERE hasil_klasifikasi = 'Tidak Laris'")
    total_tidak_laris = cursor.fetchone()["total"]
    cursor.execute(
        """
        SELECT
            p.kode_produk,
            p.nama_produk,
            p.kategori,
            SUM(pjl.jumlah_terjual) AS total_terjual,
            SUM(pjl.total_penjualan) AS total_penjualan
        FROM penjualan pjl
        JOIN produk p ON p.id = pjl.produk_id
        GROUP BY p.id, p.kode_produk, p.nama_produk, p.kategori
        ORDER BY total_terjual DESC, total_penjualan DESC
        LIMIT 5
        """
    )
    top_produk_rows = cursor.fetchall()
    cursor.execute("SELECT COALESCE(SUM(jumlah_terjual), 0) AS total FROM penjualan")
    total_produk_terjual = float(cursor.fetchone()["total"] or 0)
    produk_terlaris = top_produk_rows[0] if top_produk_rows else None
    produk_terlaris_persen = (
        (float(produk_terlaris["total_terjual"] or 0) / total_produk_terjual) * 100
        if produk_terlaris and total_produk_terjual else 0
    )
    top_produk_chart = [
        {
            "nama_produk": item["nama_produk"],
            "total_terjual": float(item["total_terjual"] or 0),
            "persentase": round((float(item["total_terjual"] or 0) / total_produk_terjual) * 100, 2)
            if total_produk_terjual else 0,
        }
        for item in top_produk_rows
    ]
    cursor.close()

    return render_template(
        "dashboard.html",
        total_produk=total_produk,
        total_penjualan=total_penjualan,
        total_klasifikasi=total_klasifikasi,
        total_laris=total_laris,
        total_tidak_laris=total_tidak_laris,
        produk_terlaris=produk_terlaris,
        produk_terlaris_persen=produk_terlaris_persen,
        total_produk_terjual=total_produk_terjual,
        top_produk_chart=top_produk_chart,
    )


@app.route("/produk")
@login_required
def produk():
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT * FROM produk ORDER BY id DESC")
    data_produk = cursor.fetchall()
    cursor.close()
    return render_template("produk.html", data_produk=data_produk)


@app.route("/produk/tambah", methods=["POST"])
@login_required
def produk_tambah():
    kode_produk = request.form.get("kode_produk", "").strip().upper()
    nama_produk = request.form.get("nama_produk", "").strip()
    kategori = request.form.get("kategori", "").strip()
    harga = request.form.get("harga", "0").strip() or 0
    stok = request.form.get("stok", "0").strip() or 0
    satuan = request.form.get("satuan", "pcs").strip() or "pcs"

    if not kode_produk or not nama_produk or not kategori:
        flash("Kode, nama produk, dan kategori wajib diisi.", "warning")
        return redirect(url_for("produk"))

    try:
        cursor = mysql.connection.cursor()
        cursor.execute(
            """
            INSERT INTO produk (kode_produk, nama_produk, kategori, harga, stok, satuan)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (kode_produk, nama_produk, kategori, harga, stok, satuan),
        )
        mysql.connection.commit()
        cursor.close()
        flash("Data produk berhasil ditambahkan.", "success")
    except Exception:
        mysql.connection.rollback()
        flash("Data produk gagal ditambahkan. Pastikan kode produk belum digunakan.", "danger")

    return redirect(url_for("produk"))


@app.route("/produk/edit/<int:produk_id>", methods=["POST"])
@login_required
def produk_edit(produk_id):
    kode_produk = request.form.get("kode_produk", "").strip().upper()
    nama_produk = request.form.get("nama_produk", "").strip()
    kategori = request.form.get("kategori", "").strip()
    harga = request.form.get("harga", "0").strip() or 0
    stok = request.form.get("stok", "0").strip() or 0
    satuan = request.form.get("satuan", "pcs").strip() or "pcs"

    if not kode_produk or not nama_produk or not kategori:
        flash("Kode, nama produk, dan kategori wajib diisi.", "warning")
        return redirect(url_for("produk"))

    try:
        cursor = mysql.connection.cursor()
        cursor.execute(
            """
            UPDATE produk
            SET kode_produk = %s, nama_produk = %s, kategori = %s, harga = %s, stok = %s, satuan = %s
            WHERE id = %s
            """,
            (kode_produk, nama_produk, kategori, harga, stok, satuan, produk_id),
        )
        mysql.connection.commit()
        cursor.close()
        flash("Data produk berhasil diperbarui.", "success")
    except Exception:
        mysql.connection.rollback()
        flash("Data produk gagal diperbarui. Pastikan kode produk tidak duplikat.", "danger")

    return redirect(url_for("produk"))


@app.route("/produk/upload", methods=["POST"])
@login_required
def produk_upload():
    uploaded_file = request.files.get("file_produk")

    if not uploaded_file or not uploaded_file.filename:
        flash("Pilih file CSV atau Excel produk terlebih dahulu.", "warning")
        return redirect(url_for("produk"))

    try:
        rows = parse_sales_rows(uploaded_file)
    except Exception as error:
        flash(str(error), "danger")
        return redirect(url_for("produk"))

    cursor = mysql.connection.cursor()
    imported = 0
    skipped = 0

    for row in rows:
        kode_produk = str(row.get("kode_produk") or "").strip().upper()
        nama_produk = str(row.get("nama_produk") or "").strip()
        kategori = str(row.get("kategori") or "").strip()
        harga = parse_float(row.get("harga") or 0)
        stok = parse_int(row.get("stok") or 0)
        satuan = str(row.get("satuan") or "pcs").strip() or "pcs"

        if not kode_produk or not nama_produk or not kategori or harga < 0 or stok < 0:
            skipped += 1
            continue

        cursor.execute(
            """
            INSERT INTO produk (kode_produk, nama_produk, kategori, harga, stok, satuan)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                nama_produk = VALUES(nama_produk),
                kategori = VALUES(kategori),
                harga = VALUES(harga),
                stok = VALUES(stok),
                satuan = VALUES(satuan)
            """,
            (kode_produk, nama_produk, kategori, harga, stok, satuan),
        )
        imported += 1

    mysql.connection.commit()
    cursor.close()
    flash(f"Upload produk selesai. {imported} data berhasil disimpan, {skipped} data dilewati.", "success")
    return redirect(url_for("produk"))


@app.route("/produk/template")
@login_required
def produk_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Format Upload Produk"

    headers = ["kode_produk", "nama_produk", "kategori", "harga", "stok", "satuan"]
    sheet.append(headers)
    sheet.append(["PRD001", "Facial Wash SR12", "Facial Wash", 35000, 50, "pcs"])
    sheet.append(["PRD002", "Serum Brightening SR12", "Serum", 75000, 30, "pcs"])

    header_fill = PatternFill("solid", fgColor="F9D7E8")
    header_font = Font(bold=True, color="75445C")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = [16, 32, 22, 14, 12, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    sheet["H1"] = "Catatan"
    sheet["H2"] = "kode_produk wajib unik."
    sheet["H3"] = "Jika kode_produk sudah ada, data produk akan diperbarui."
    sheet["H4"] = "harga dan stok wajib angka."
    sheet["H5"] = "satuan boleh diisi pcs/botol/box sesuai kebutuhan."
    sheet.column_dimensions["H"].width = 62

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=format_upload_data_produk.xlsx"},
    )


@app.route("/produk/hapus/<int:produk_id>", methods=["POST"])
@login_required
@admin_required
def produk_hapus(produk_id):
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM penjualan WHERE produk_id = %s", (produk_id,))
    total_dipakai = cursor.fetchone()["total"]

    if total_dipakai > 0:
        cursor.close()
        flash("Produk tidak dapat dihapus karena sudah digunakan pada data penjualan.", "warning")
        return redirect(url_for("produk"))

    cursor.execute("DELETE FROM produk WHERE id = %s", (produk_id,))
    mysql.connection.commit()
    cursor.close()
    flash("Data produk berhasil dihapus.", "success")
    return redirect(url_for("produk"))


@app.route("/penjualan")
@login_required
def penjualan():
    cursor = mysql.connection.cursor()
    cursor.execute(
        """
        SELECT pjl.*, p.nama_produk, p.kode_produk, p.kategori, p.harga
        FROM penjualan pjl
        JOIN produk p ON p.id = pjl.produk_id
        ORDER BY pjl.id DESC
        """
    )
    data_penjualan = cursor.fetchall()
    cursor.execute("SELECT id, kode_produk, nama_produk, kategori, harga, stok FROM produk ORDER BY nama_produk ASC")
    data_produk = cursor.fetchall()
    cursor.close()
    return render_template("penjualan.html", data_penjualan=data_penjualan, data_produk=data_produk)


@app.route("/penjualan/tambah", methods=["POST"])
@login_required
def penjualan_tambah():
    produk_id = request.form.get("produk_id", "").strip()
    tanggal_penjualan = normalize_date_input(request.form.get("tanggal_penjualan", ""))
    bulan = request.form.get("bulan", "").strip()
    tahun = request.form.get("tahun", "").strip()
    jumlah_terjual = parse_int(request.form.get("jumlah_terjual", "0"))

    if not produk_id or not tanggal_penjualan or not bulan or not tahun:
        flash("Produk, tanggal, bulan, dan tahun wajib diisi.", "warning")
        return redirect(url_for("penjualan"))

    cursor = mysql.connection.cursor()
    cursor.execute("SELECT harga FROM produk WHERE id = %s", (produk_id,))
    produk_data = cursor.fetchone()

    if not produk_data:
        cursor.close()
        flash("Produk tidak ditemukan.", "danger")
        return redirect(url_for("penjualan"))

    total_penjualan = float(produk_data["harga"]) * jumlah_terjual
    cursor.execute(
        """
        INSERT INTO penjualan
        (produk_id, tanggal_penjualan, bulan, tahun, jumlah_terjual, total_penjualan)
        VALUES (%s, %s, %s, %s, %s, %s)
        """,
        (produk_id, tanggal_penjualan, bulan, tahun, jumlah_terjual, total_penjualan),
    )
    mysql.connection.commit()
    cursor.close()
    flash("Data penjualan berhasil ditambahkan.", "success")
    return redirect(url_for("penjualan"))


@app.route("/penjualan/edit/<int:penjualan_id>", methods=["POST"])
@login_required
def penjualan_edit(penjualan_id):
    produk_id = request.form.get("produk_id", "").strip()
    tanggal_penjualan = normalize_date_input(request.form.get("tanggal_penjualan", ""))
    bulan = request.form.get("bulan", "").strip()
    tahun = request.form.get("tahun", "").strip()
    jumlah_terjual = parse_int(request.form.get("jumlah_terjual", "0"))

    if not produk_id or not tanggal_penjualan or not bulan or not tahun:
        flash("Produk, tanggal, bulan, dan tahun wajib diisi.", "warning")
        return redirect(url_for("penjualan"))

    cursor = mysql.connection.cursor()
    cursor.execute("SELECT harga FROM produk WHERE id = %s", (produk_id,))
    produk_data = cursor.fetchone()

    if not produk_data:
        cursor.close()
        flash("Produk tidak ditemukan.", "danger")
        return redirect(url_for("penjualan"))

    total_penjualan = float(produk_data["harga"]) * jumlah_terjual
    cursor.execute(
        """
        UPDATE penjualan
        SET produk_id = %s, tanggal_penjualan = %s, bulan = %s, tahun = %s,
            jumlah_terjual = %s, total_penjualan = %s
        WHERE id = %s
        """,
        (
            produk_id,
            tanggal_penjualan,
            bulan,
            tahun,
            jumlah_terjual,
            total_penjualan,
            penjualan_id,
        ),
    )
    mysql.connection.commit()
    cursor.close()
    flash("Data penjualan berhasil diperbarui.", "success")
    return redirect(url_for("penjualan"))


@app.route("/penjualan/upload", methods=["POST"])
@login_required
def penjualan_upload():
    uploaded_file = request.files.get("file_penjualan")

    if not uploaded_file or not uploaded_file.filename:
        flash("Pilih file CSV atau Excel terlebih dahulu.", "warning")
        return redirect(url_for("penjualan"))

    try:
        rows = parse_sales_rows(uploaded_file)
    except Exception as error:
        flash(str(error), "danger")
        return redirect(url_for("penjualan"))

    cursor = mysql.connection.cursor()
    imported = 0
    skipped = 0

    for row in rows:
        kode_produk = str(row.get("kode_produk") or "").strip().upper()
        nama_produk = str(row.get("nama_produk") or row.get("produk") or "").strip()
        kategori = str(row.get("kategori") or "").strip()
        product_key = kode_produk or nama_produk
        raw_date = row.get("tanggal") or row.get("tanggal_penjualan") or ""
        tanggal = normalize_date_input(raw_date)
        bulan = str(row.get("bulan") or "").strip()
        tahun = str(row.get("tahun") or "").strip()
        jumlah_terjual = parse_int(row.get("jumlah_terjual") or row.get("terjual") or 0)
        harga_file = parse_float(row.get("harga") or 0)
        total_file = parse_float(row.get("total_penjualan") or 0)

        if not product_key or not tanggal or jumlah_terjual < 0:
            skipped += 1
            continue

        cursor.execute(
            """
            SELECT id, harga
            FROM produk
            WHERE kode_produk = %s OR nama_produk = %s
            LIMIT 1
            """,
            (product_key, product_key),
        )
        produk_data = cursor.fetchone()

        if not produk_data:
            if not kode_produk or not nama_produk or not kategori or harga_file <= 0:
                skipped += 1
                continue
            cursor.execute(
                """
                INSERT INTO produk (kode_produk, nama_produk, kategori, harga, stok, satuan)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (kode_produk, nama_produk, kategori, harga_file, 0, "pcs"),
            )
            produk_id = cursor.lastrowid
            produk_harga = harga_file
        else:
            produk_id = produk_data["id"]
            produk_harga = float(produk_data["harga"])

        parsed_date = datetime.strptime(tanggal, "%Y-%m-%d")
        bulan = bulan or month_name_id(parsed_date.month)
        tahun = tahun or str(parsed_date.year)
        total_penjualan = total_file if total_file > 0 else produk_harga * jumlah_terjual

        cursor.execute(
            """
            INSERT INTO penjualan
            (produk_id, tanggal_penjualan, bulan, tahun, jumlah_terjual, total_penjualan)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (produk_id, tanggal, bulan, tahun, jumlah_terjual, total_penjualan),
        )
        imported += 1

    mysql.connection.commit()
    cursor.close()
    flash(f"Upload selesai. {imported} data berhasil diimport, {skipped} data dilewati.", "success")
    return redirect(url_for("penjualan"))


@app.route("/penjualan/template")
@login_required
def penjualan_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Format Upload Penjualan"

    headers = [
        "tanggal_penjualan",
        "kode_produk",
        "nama_produk",
        "kategori",
        "jumlah_terjual",
        "harga",
        "total_penjualan",
    ]
    sheet.append(headers)
    sheet.append(["12/06/2026", "PRD001", "Facial Wash SR12", "Facial Wash", 10, 35000, ""])
    sheet.append(["13/06/2026", "PRD002", "Serum Brightening SR12", "Serum", 5, 75000, 375000])

    header_fill = PatternFill("solid", fgColor="F9D7E8")
    header_font = Font(bold=True, color="75445C")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = [20, 16, 30, 20, 18, 14, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    sheet["I1"] = "Catatan"
    sheet["I2"] = "tanggal_penjualan wajib format dd/mm/yyyy."
    sheet["I3"] = "kode_produk harus sama dengan Data Produk jika produk sudah ada."
    sheet["I4"] = "Jika produk belum ada, isi kode_produk, nama_produk, kategori, dan harga."
    sheet["I5"] = "total_penjualan boleh dikosongkan, sistem akan menghitung harga x jumlah_terjual."
    sheet.column_dimensions["I"].width = 72

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=format_upload_data_penjualan.xlsx"},
    )


@app.route("/penjualan/hapus/<int:penjualan_id>", methods=["POST"])
@login_required
@admin_required
def penjualan_hapus(penjualan_id):
    cursor = mysql.connection.cursor()
    cursor.execute("DELETE FROM penjualan WHERE id = %s", (penjualan_id,))
    mysql.connection.commit()
    cursor.close()
    flash("Data penjualan berhasil dihapus.", "success")
    return redirect(url_for("penjualan"))


@app.route("/dataset")
@login_required
def dataset():
    flash("Dataset dibentuk otomatis saat Proses C4.5 dijalankan.", "info")
    return redirect(url_for("proses_c45"))


@app.route("/proses-c45", methods=["GET", "POST"])
@login_required
@admin_required
def proses_c45():
    tanggal_awal, tanggal_akhir = get_period_inputs()
    raw_dataset = get_aggregated_sales(tanggal_awal, tanggal_akhir) if tanggal_awal and tanggal_akhir else []

    if tanggal_awal and tanggal_akhir and tanggal_awal > tanggal_akhir:
        flash("Tanggal awal tidak boleh lebih besar dari tanggal akhir.", "warning")
        return redirect(url_for("proses_c45"))

    classification_rows, dataset, base_entropy, calculations, best_attribute = build_classification_rows(raw_dataset)
    class_counts = {}

    for row in dataset:
        label = row.get("label_klasifikasi") or "-"
        class_counts[label] = class_counts.get(label, 0) + 1

    rules = build_simple_rules(dataset, best_attribute["attribute"]) if best_attribute else []
    total_terjual = sum(float(row.get("jumlah_terjual") or 0) for row in raw_dataset)
    total_penjualan = sum(float(row.get("total_penjualan") or 0) for row in raw_dataset)
    average_sold = classification_rows[0]["rata_rata_penjualan"] if classification_rows else 0

    if request.method == "POST":
        if not tanggal_awal or not tanggal_akhir:
            flash("Tanggal awal dan tanggal akhir wajib diisi.", "warning")
            return redirect(url_for("proses_c45"))
        if not raw_dataset:
            flash("Tidak ada data penjualan pada periode yang dipilih.", "warning")
            return redirect(url_for("proses_c45", tanggal_awal=format_date_id(tanggal_awal), tanggal_akhir=format_date_id(tanggal_akhir)))

        parsed_start = datetime.strptime(tanggal_awal, "%Y-%m-%d")
        bulan = month_name_id(parsed_start.month)
        tahun = str(parsed_start.year)
        cursor = mysql.connection.cursor()
        cursor.execute(
            "DELETE FROM hasil_klasifikasi WHERE tanggal_awal = %s AND tanggal_akhir = %s",
            (tanggal_awal, tanggal_akhir),
        )

        for row in classification_rows:
            cursor.execute(
                """
                INSERT INTO hasil_klasifikasi
                (produk_id, kode_produk, nama_produk, kategori, tanggal_awal, tanggal_akhir,
                 bulan, tahun, jumlah_terjual, total_penjualan, rata_rata_penjualan,
                 hasil_klasifikasi, keterangan)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    row["produk_id"],
                    row["kode_produk"],
                    row["nama_produk"],
                    row["kategori"],
                    tanggal_awal,
                    tanggal_akhir,
                    bulan,
                    tahun,
                    row["jumlah_terjual"],
                    row["total_penjualan"],
                    row["rata_rata_penjualan"],
                    row["hasil_klasifikasi"],
                    row["keterangan"],
                ),
            )

        mysql.connection.commit()
        cursor.close()
        flash("Proses C4.5 berhasil dijalankan dan hasil klasifikasi disimpan.", "success")
        return redirect(url_for("klasifikasi", tanggal_awal=format_date_id(tanggal_awal), tanggal_akhir=format_date_id(tanggal_akhir)))

    return render_template(
        "proses_c45.html",
        total_dataset=len(dataset),
        raw_dataset=raw_dataset,
        classification_rows=classification_rows,
        class_counts=class_counts,
        base_entropy=base_entropy,
        calculations=calculations,
        best_attribute=best_attribute,
        rules=rules,
        tanggal_awal=format_date_id(tanggal_awal) if tanggal_awal else "",
        tanggal_akhir=format_date_id(tanggal_akhir) if tanggal_akhir else "",
        total_terjual=total_terjual,
        total_penjualan=total_penjualan,
        average_sold=average_sold,
    )


@app.route("/klasifikasi")
@login_required
def klasifikasi():
    hasil, filters = fetch_classification_results()
    return render_template("klasifikasi.html", hasil=hasil, filters=filters)


@app.route("/klasifikasi/proses", methods=["POST"])
@login_required
@admin_required
def klasifikasi_proses():
    flash("Proses klasifikasi sekarang dijalankan melalui menu Proses C4.5 dengan periode tanggal.", "info")
    return redirect(url_for("proses_c45"))


@app.route("/laporan")
@login_required
def laporan():
    data_laporan, filters = fetch_classification_results()
    return render_template("laporan.html", data_laporan=data_laporan, filters=filters)


@app.route("/laporan/download")
@login_required
def laporan_download():
    data_laporan, _ = fetch_classification_results()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Klasifikasi"
    sheet.append([
        "No",
        "Periode",
        "Kode Produk",
        "Nama Produk",
        "Kategori",
        "Jumlah Terjual",
        "Total Penjualan",
        "Hasil Klasifikasi",
        "Keterangan",
    ])

    for index, item in enumerate(data_laporan, start=1):
        sheet.append([
            index,
            format_period(item["tanggal_awal"], item["tanggal_akhir"]),
            item["kode_produk"] or "-",
            item["nama_produk"],
            item["kategori"],
            item["jumlah_terjual"],
            float(item["total_penjualan"] or 0),
            item["hasil_klasifikasi"],
            item["keterangan"] or "-",
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=laporan_klasifikasi_c45.xlsx"},
    )


@app.route("/users")
@login_required
@admin_required
def users():
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT id, nama, username, role, created_at FROM users ORDER BY id DESC")
    data_users = cursor.fetchall()
    cursor.close()
    return render_template("users.html", data_users=data_users)


@app.route("/users/tambah", methods=["POST"])
@login_required
@admin_required
def user_tambah():
    nama = request.form.get("nama", "").strip()
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "staff").strip()

    if not nama or not username or not password or role not in ("admin", "staff"):
        flash("Nama, username, password, dan role wajib diisi.", "warning")
        return redirect(url_for("users"))

    try:
        cursor = mysql.connection.cursor()
        cursor.execute(
            """
            INSERT INTO users (nama, username, password, role)
            VALUES (%s, %s, %s, %s)
            """,
            (nama, username, generate_password_hash(password), role),
        )
        mysql.connection.commit()
        cursor.close()
        flash("Data user berhasil ditambahkan.", "success")
    except Exception:
        mysql.connection.rollback()
        flash("Data user gagal ditambahkan. Pastikan username belum digunakan.", "danger")

    return redirect(url_for("users"))


@app.route("/users/edit/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def user_edit(user_id):
    nama = request.form.get("nama", "").strip()
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "staff").strip()

    if not nama or not username or role not in ("admin", "staff"):
        flash("Nama, username, dan role wajib diisi.", "warning")
        return redirect(url_for("users"))

    if user_id == session.get("user_id") and role != "admin":
        flash("Akun yang sedang login tidak boleh mengubah role sendiri menjadi Staff.", "warning")
        return redirect(url_for("users"))

    try:
        cursor = mysql.connection.cursor()
        if password:
            cursor.execute(
                """
                UPDATE users
                SET nama = %s, username = %s, password = %s, role = %s
                WHERE id = %s
                """,
                (nama, username, generate_password_hash(password), role, user_id),
            )
        else:
            cursor.execute(
                """
                UPDATE users
                SET nama = %s, username = %s, role = %s
                WHERE id = %s
                """,
                (nama, username, role, user_id),
            )

        mysql.connection.commit()
        cursor.close()

        if user_id == session.get("user_id"):
            session["nama"] = nama
            session["role"] = role

        flash("Data user berhasil diperbarui.", "success")
    except Exception:
        mysql.connection.rollback()
        flash("Data user gagal diperbarui. Pastikan username tidak duplikat.", "danger")

    return redirect(url_for("users"))


@app.route("/users/hapus/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def user_hapus(user_id):
    if user_id == session.get("user_id"):
        flash("Akun yang sedang login tidak dapat dihapus.", "warning")
        return redirect(url_for("users"))

    cursor = mysql.connection.cursor()
    cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
    mysql.connection.commit()
    cursor.close()
    flash("Data user berhasil dihapus.", "success")
    return redirect(url_for("users"))


if __name__ == "__main__":
    app.run(debug=True)
